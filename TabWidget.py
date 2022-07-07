from PyQt5 import uic
from PyQt5.QtCore import Qt, QRegExp
from PyQt5.QtGui import QDoubleValidator, QRegExpValidator
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QWidget, QListWidgetItem
from openpyxl import load_workbook

from analysis.Append2Excel import append_df_to_excel
from analysis.JetTestCases import JetTestCases
from checking_data import *
from utils import get_dataframe


class TabWidget(QWidget):
    filenames = []
    dataframe = ""
    suspicious_words = []
    dismissed_employee = []
    patterns = []
    estimates = []
    gl_columns = []

    def __init__(self, filenames):
        super(TabWidget, self).__init__()
        # self.setWindowFlags(Qt.Window | Qt.WindowStaysOnTopHint)
        uic.loadUi('user_interface/tabWidget.ui', self)
        # self.ui = uic.loadUi('user_interface/tabWidget.ui', self)
        self.tabWidget.setCurrentIndex(0)
        self.cm_ctn_holiday.addItem('Canada', ['Alberta', 'Ontario', 'Quebec'])  # index 0
        self.cm_ctn_holiday.addItem('USA', ['California', 'Florida', 'Illinois'])  # index 1
        self.cm_ctn_holiday.currentIndexChanged.connect(self.updateCityCombo)
        self.updateCityCombo(self.cm_ctn_holiday.currentIndex())

        self.dataframe = pd.read_excel(filenames[0], sheet_name=0, index_col=None)

        self.gl_columns = self.dataframe.columns.values.tolist()
        self.filenames = filenames

        self.disable_inputs()
        self.predefined_words()

        # self.initial_testing()
        # self.init_listview()
        # self.init_listWidget()

        self.oob_cb.clicked.connect(self.check_state)
        self.wk_cb.clicked.connect(self.check_state)
        self.hl_cb.clicked.connect(self.check_state)
        self.unt_cb.clicked.connect(self.check_state)
        self.bfd_cb.clicked.connect(self.check_state)
        self.sus_cb.clicked.connect(self.check_state)
        self.ov_cb.clicked.connect(self.check_state)
        self.us_cb.clicked.connect(self.check_state)
        self.ta_cb.clicked.connect(self.check_state)
        self.wh_cb.clicked.connect(self.check_state)
        self.ac_cb.clicked.connect(self.check_state)
        self.sq_cb.clicked.connect(self.check_state)
        self.dp_cb.clicked.connect(self.check_state)
        self.sp_cb.clicked.connect(self.check_state)
        self.nd_cb.clicked.connect(self.check_state)
        self.wa_cb.clicked.connect(self.check_state)

        self.sus_pb_add.clicked.connect(self.add_words)
        self.sus_pb_remove.clicked.connect(self.remove_words)
        self.sus_pb_clear.clicked.connect(self.clear_words)
        self.sus_pb_move.clicked.connect(self.move_word)
        self.sus_pb_move_back.clicked.connect(self.move_back_word)

        self.wh_pb_add.clicked.connect(self.add_pattern)
        self.wh_pb_remove.clicked.connect(self.remove_pattern)
        self.wh_pb_clear.clicked.connect(self.clear_patterns)

        self.ac_pb_add.clicked.connect(self.add_estimates)
        self.ac_pb_remove.clicked.connect(self.remove_estimates)
        self.ac_pb_clear.clicked.connect(self.clear_estimates)

        self.ta_le_deviation.valueChanged.connect(self.valuechange)
        self.ta_le_deviation.editingFinished.connect(lambda: self.valuechange())

        self.pb_prev.clicked.connect(self.prev_tab)
        self.pb_next.clicked.connect(self.next_tab)

        self.usan_pb_refresh.clicked.connect(self.refresh_names)

        self.pb_runtests.clicked.connect(self.run_all_test)

        self.init_column_names()
        self.init_duplicate_listwidget()

    def init_column_names(self):
        self.cmb_out_of_bond.addItems(self.gl_columns)
        self.cmb_weekend.addItems(self.gl_columns)
        self.cmb_holiday.addItems(self.gl_columns)
        self.cmb_unusual.addItems(self.gl_columns)
        self.cmb_forward_date.addItems(self.gl_columns)
        self.cmb_back_date.addItems(self.gl_columns)
        self.cmb_suspicious.addItems(self.gl_columns)
        self.cmb_over_scope.addItems(self.gl_columns)
        self.cmb_user_analysis.addItems(self.gl_columns)
        self.cmb_treshold.addItems(self.gl_columns)
        self.cmb_whole.addItems(self.gl_columns)
        self.cmb_estimates.addItems(self.gl_columns)

        self.cmb_sequential.addItems(self.gl_columns)
        # self.cmb_duplicates.addItems(self.gl_columns)
        self.cmb_suspense.addItems(self.gl_columns)
        self.cmb_nodescription.addItems(self.gl_columns)
        self.cmb_nodescription_date.addItems(self.gl_columns)
        self.cmb_word.addItems(self.gl_columns)

        self.cmb_out_of_bond.setCurrentIndex(0)
        self.cmb_weekend.setCurrentIndex(0)
        self.cmb_holiday.setCurrentIndex(0)
        self.cmb_unusual.setCurrentIndex(0)
        self.cmb_forward_date.setCurrentIndex(0)
        self.cmb_back_date.setCurrentIndex(0)
        self.cmb_suspicious.setCurrentIndex(0)
        self.cmb_over_scope.setCurrentIndex(0)
        self.cmb_user_analysis.setCurrentIndex(0)
        self.cmb_treshold.setCurrentIndex(0)
        self.cmb_whole.setCurrentIndex(0)
        self.cmb_estimates.setCurrentIndex(0)

        self.cmb_sequential.setCurrentIndex(0)
        # self.cmb_duplicates.setCurrentIndex(0)
        self.cmb_suspense.setCurrentIndex(0)
        self.cmb_nodescription.setCurrentIndex(0)
        self.cmb_nodescription_date.setCurrentIndex(0)
        self.cmb_word.setCurrentIndex(0)

    def valuechange(self):
        treshold = int(self.ta_le_treshold.text())
        deviation = int(self.ta_le_deviation.text())
        if treshold - deviation < 0:
            info_message(self,
                         text=f"Deviation must be less than treshold")
            self.ta_le_deviation.setValue(0)
        # self.l1.setText("current value:" + str(self.sp.value()))

    def updateCityCombo(self, index):
        self.cm_pr_holiday.clear()
        cities = self.cm_ctn_holiday.itemData(index)
        if cities:
            self.cm_pr_holiday.addItems(cities)

    # def closeEvent(self, event):
    #     reply = QMessageBox.question(self, 'Close?',
    #                                  'Are you sure you want to close?',
    #                                  QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    #
    #     if reply == QMessageBox.Yes:
    #         if not type(event) == bool:
    #             event.accept()
    #         else:
    #             self.close()
    #     else:
    #         if not type(event) == bool:
    #             event.ignore()
    # self.parent().show()

    def check_state(self):
        if self.oob_cb.isChecked():
            self.fr_of.setEnabled(True)
        else:
            self.fr_of.setDisabled(True)

        if self.wk_cb.isChecked():
            self.fr_wk.setEnabled(True)
        else:
            self.fr_wk.setDisabled(True)

        if self.hl_cb.isChecked():
            self.fr_hl.setEnabled(True)
        else:
            self.fr_hl.setDisabled(True)

        if self.unt_cb.isChecked():
            self.fr_unt.setEnabled(True)
        else:
            self.fr_unt.setDisabled(True)

        if self.bfd_cb.isChecked():
            self.fr_back.setEnabled(True)
        else:
            self.fr_back.setDisabled(True)

        if self.sus_cb.isChecked():
            self.fr_sus.setEnabled(True)
        else:
            self.fr_sus.setDisabled(True)

        if self.ov_cb.isChecked():
            self.fr_treshold1.setEnabled(True)
        else:
            self.fr_treshold1.setDisabled(True)

        if self.us_cb.isChecked():
            self.fr_ua.setEnabled(True)
        else:
            self.fr_ua.setDisabled(True)

        if self.ta_cb.isChecked():
            self.fr_treshold2.setEnabled(True)
        else:
            self.fr_treshold2.setDisabled(True)

        if self.wh_cb.isChecked():
            self.fr_whole.setEnabled(True)
        else:
            self.fr_whole.setDisabled(True)

        if self.ac_cb.isChecked():
            self.fr_estimates.setEnabled(True)
        else:
            self.fr_estimates.setDisabled(True)

        if self.sq_cb.isChecked():
            self.fr_sequential.setEnabled(True)
        else:
            self.fr_sequential.setDisabled(True)

        if self.dp_cb.isChecked():
            self.fr_duplicates.setEnabled(True)
        else:
            self.fr_duplicates.setDisabled(True)

        if self.sp_cb.isChecked():
            self.fr_suspense.setEnabled(True)
        else:
            self.fr_suspense.setDisabled(True)

        if self.nd_cb.isChecked():
            self.fr_nodescription.setEnabled(True)
        else:
            self.fr_nodescription.setDisabled(True)

        if self.wa_cb.isChecked():
            self.fr_word.setEnabled(True)
        else:
            self.fr_word.setDisabled(True)

    def disable_inputs(self):
        self.fr_of.setDisabled(True)
        self.fr_wk.setDisabled(True)
        self.fr_hl.setDisabled(True)
        self.fr_unt.setDisabled(True)
        self.fr_back.setDisabled(True)
        self.fr_sus.setDisabled(True)
        self.fr_treshold1.setDisabled(True)
        self.fr_ua.setDisabled(True)
        self.fr_treshold2.setDisabled(True)
        self.fr_whole.setDisabled(True)
        self.fr_estimates.setDisabled(True)
        self.fr_sequential.setDisabled(True)
        self.fr_duplicates.setDisabled(True)
        self.fr_suspense.setDisabled(True)
        self.fr_nodescription.setDisabled(True)
        self.fr_word.setDisabled(True)

    def prev_tab(self):
        index = self.tabWidget.currentIndex()
        print(index)
        if index != 0:
            self.tabWidget.setCurrentIndex(int(index - 1))

    def next_tab(self):
        index = self.tabWidget.currentIndex()
        print(index)
        if index < 9:
            self.tabWidget.setCurrentIndex(int(index + 1))

    def run_all_test(self):
        file_index = 1
        failed_tests = []
        row = 2
        FINAL_RESULT_FILE = copy_file_from_src_to_dest()
        for file in self.filenames:
            k = 1
            print(f'{k} ----------------------------------------------->')
            file_name = file.split('/')[-1][:-5]
            folder_name = file_name
            # folder = file.split('/')[-2]
            # print(folder)
            dataframe = pd.read_excel(file, sheet_name=0, index_col=None)
            jet_test = JetTestCases(dataframe)
            print(f'{file_index} -> ', folder_name, ' is running')
            file_index += 1
            # check_path(f'results/{folder_name}')

            if self.oob_cb.isChecked():
                date_col = self.cmb_out_of_bond.currentText()
                start_datetime = self.oob_dte_start.dateTime().toPyDateTime()
                end_datetime = self.oob_dte_end.dateTime().toPyDateTime()
                print(start_datetime, end_datetime)
                df, is_passed = jet_test.get_out_of_bound_entries(date=date_col, start=start_datetime, end=end_datetime)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Out of bound entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Out of bound entries is failed, Error!')
                    failed_tests.append('Out of bound entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Out of bound entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.wk_cb.isChecked():
                date_col = self.cmb_weekend.currentText()
                weekends = [
                    self.week_day(self.wk_chb_monday),
                    self.week_day(self.wk_chb_tuesday),
                    self.week_day(self.wk_chb_wednesday),
                    self.week_day(self.wk_chb_thursday),
                    self.week_day(self.wk_chb_friday),
                    self.week_day(self.wk_chb_saturday),
                    self.week_day(self.wk_chb_sunday)
                ]
                weekends = [i for i, x in enumerate(weekends) if x == 1]
                df, is_passed = jet_test.get_weekend_entries(date=date_col, week_days=weekends)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Weekend entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Weekend entries is failed, Error!')
                    failed_tests.append('Weekend entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Weekend entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.hl_cb.isChecked():
                date_col = self.cmb_holiday.currentText()
                countries = ["CA", "US"]
                PROVINCES = ["AB", "ON", "QC"]
                STATES = ["CA", "FL", "IL"]
                print(self.cm_ctn_holiday.currentIndex())
                cntry = countries[self.cm_ctn_holiday.currentIndex()]
                print(cntry)
                if cntry == 'CA':
                    prov = PROVINCES[self.cm_pr_holiday.currentIndex()]
                else:
                    prov = STATES[self.cm_pr_holiday.currentIndex()]
                df, is_passed = jet_test.get_holiday_entries(date=date_col, country=cntry, province=prov)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Holiday entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Holiday entries is failed, Error!')
                    failed_tests.append('Holiday entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Holiday entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.unt_cb.isChecked():
                date_col = self.cmb_unusual.currentText()
                finish = self.unt_finish_time.time().toString()
                print(finish)
                df, is_passed = jet_test.get_unusual_times(date=date_col, finish_time=finish)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Unusual times')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Unusual times is failed, Error!')
                    failed_tests.append('Unusual times')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Unusual times is empty')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.bfd_cb.isChecked():
                create_d = self.cmb_back_date.currentText()
                posted_d = self.cmb_forward_date.currentText()
                days = int(self.bfd_le_days.text())
                df, is_passed = jet_test.get_back_forward_date_entries(
                    days=days,
                    created_date=create_d,
                    posted_date=posted_d
                )
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Back - Forward date entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Back - Forward date entries is failed, Error!')
                    failed_tests.append('Back - Forward date entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Back - Forward date entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.sus_cb.isChecked():
                descr = self.cmb_suspicious.currentText()
                items = set()
                for x in range(self.sus_lwidget.count()):
                    items.add(self.sus_lwidget.item(x).text())
                self.suspicious_words = list(items)
                # print(self.suspicious_words)
                df, is_passed = jet_test.get_suspicious_description(description=descr, words=self.suspicious_words)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspicious description')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspicious description is failed, Error!')
                    failed_tests.append('Suspicious description')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspicious description is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.ov_cb.isChecked():
                over_scope = int(self.ov_le_treshold.text())
                amount_col = self.cmb_over_scope.currentText()
                print(over_scope)
                df, is_passed = jet_test.get_over_scope_entries(amount=amount_col, given_amount=over_scope)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Over scope entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Over scope entries is failed, Error!')
                    failed_tests.append('Over scope entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Over scope entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.us_cb.isChecked():
                employee_col = self.cmb_user_analysis.currentText()
                names = []
                for i in range(self.user_listwidget.count() - 1):
                    item = self.user_listwidget.item(i)
                    if item.checkState():
                        names.append(item.text())
                print(names)
                df, is_passed = jet_test.get_employee_by_name(created_by=employee_col, names=names)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'User analysis')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'User analysis is failed, Error!')
                    failed_tests.append('User analysis')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'User analysis is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.ta_cb.isChecked():
                amnt = self.cmb_treshold.currentText()
                treshold = int(self.ta_le_treshold.text())
                deviation = int(self.ta_le_deviation.text())
                print(treshold, deviation)
                df, is_passed = jet_test.get_amount_between_range(amount=amnt, treshold=treshold, deviation=deviation)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Treshold analysis')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Treshold analysis is failed, Error!')
                    failed_tests.append('Treshold analysis')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Treshold analysis is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.wh_cb.isChecked():
                amount_col = self.cmb_whole.currentText()
                items = set()
                for x in range(self.wh_lwidget.count()):
                    items.add(self.wh_lwidget.item(x).text())
                self.patterns = list(items)
                # for i in range(len(self.patterns)):
                #     if self.patterns[i].find('.') != -1:
                #         ind = self.patterns[i].find('.')
                #         self.patterns[i] = self.patterns[i][:ind] + self.patterns[i][ind:]
                #     else:
                #         self.patterns[i] += '.00'
                #         print('Not found')
                print(self.patterns)
                df, is_passed = jet_test.get_whole_amounts(amount=amount_col, patterns=self.patterns)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Whole amounts')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Whole amounts is failed, Error!')
                    failed_tests.append('Whole amounts')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Whole amounts is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.ac_cb.isChecked():
                estm = self.cmb_estimates.currentText()
                items = set()
                for x in range(self.ac_lwidget.count()):
                    items.add(int(self.ac_lwidget.item(x).text()))
                self.estimates = list(items)
                print(self.estimates)
                df, is_passed = jet_test.get_account_estimates(estimates=estm, list=self.estimates)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Account estimates')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Account estimates is failed, Error!')
                    failed_tests.append('Account estimates')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Account estimates is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.sq_cb.isChecked():
                journal_num_col = self.cmb_sequential.currentText()
                df, is_passed = jet_test.get_omitted_sequential_je(journal_num=journal_num_col)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Sequential journal entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Sequential journal entries is failed, Error!')
                    failed_tests.append('Sequential journal entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Sequential journal entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.dp_cb.isChecked():
                # print(self.duplicate_listwidget.selectedIndexes())
                duplicate_col_names = []
                for i in range(self.duplicate_listwidget.count() - 1):
                    item = self.duplicate_listwidget.item(i)
                    if item.checkState():
                        duplicate_col_names.append(item.text())
                print(duplicate_col_names)
                df, is_passed = jet_test.get_duplicate_amounts(list=duplicate_col_names)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Duplicate entries')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Duplicate entries is failed, Error!')
                    failed_tests.append('Duplicate entries')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Duplicate entries is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.sp_cb.isChecked():
                string_col = self.cmb_suspense.currentText()
                df, is_passed = jet_test.get_suspensed_accounts(text_col=string_col)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspended accounts')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspended accounts is failed, Error!')
                    failed_tests.append('Suspended accounts')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Suspended accounts is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.nd_cb.isChecked():
                date_col = self.cmb_nodescription_date.currentText()
                text_col = self.cmb_nodescription.currentText()
                dt = self.dte_start_description.dateTime().toPyDateTime()
                df, is_passed = jet_test.get_no_description(date=date_col, description=text_col, chosen_date=dt)
                if not df.empty and is_passed:
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'No description in given date')
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'No description in given date is failed, Error!')
                    failed_tests.append('No description in given date')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'No description in given date is empty!')
                print("Process finished")
                k += 1
                print(f'{k} ----------------------------------------------->')
                print(df)

            if self.wa_cb.isChecked():
                descr = self.cmb_word.currentText()
                df, is_passed = jet_test.get_word_freq(description=descr)
                # print(df)
                if not df.empty and is_passed:
                    kwarg = {'folder': folder_name, 'file': str(k) + '_' + file_name,
                             'categories': df['Words'].values.tolist(), 'values': df['Words count'].values.tolist(),
                             'cell': 'D' + str(row)}
                    save_to_xlsx_file(folder_name, str(k) + '_' + file_name, df)
                    append_df_to_excel(FINAL_RESULT_FILE, df, sheet_name=str(k), index=False, header=None, kw=kwarg)
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Word analysis')
                    plot_data_to_xlsx(file_name=FINAL_RESULT_FILE, sheet_name=str(k), kw=kwarg)
                elif not is_passed:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Word analysis is failed, Error!')
                    failed_tests.append('Word analysis')
                else:
                    self.write_test_name(FINAL_RESULT_FILE, str(k), 'Word analysis is empty')
                print("Process finished")
                k += 1
            row += 20
            del jet_test
            print(folder_name + ' is processed!')
        t = 'All test are done!\nFailed jet test given below:\n'
        failed_tests.insert(0, f'Failed tests count : {len(failed_tests)}')
        info_message(self, text=t + "\n".join(failed_tests))
        self.close()

    def write_test_name(self, file_name, sheet_name, test_name):
        srcfile = load_workbook(file_name)
        srcfile[sheet_name].cell(column=1, row=1, value=test_name)
        # get sheetname from the file
        # sheets = srcfile.sheetnames
        # print(sheets)

        # To work with the first sheet (by name)
        # ws = srcfile[sheets[0]]
        # print(ws['A1'].value)

        # To work with the active sheet
        # ws = srcfile.active
        # print(ws['A1'].value)
        # sheetname = srcfile.get_sheet_by_name(sheet_name)
        # sheetname = srcfile.get_index(1+int(sheet_name))
        # write something in B2 cell of the supplied sheet
        # sheetname['A1'] = test_name
        # write to row 1,col 1 explicitly, this type of writing is useful to
        # write something in loops
        # sheetname.cell(row=1, column=1).value = 'something'

        # save it as a new file, the original file is untouched and here I am saving
        # it as xlsm(m here denotes macros).
        srcfile.save(file_name)
        srcfile.close()

    def week_day(self, day):
        return 1 if day.isChecked() else 0

    def predefined_words(self):
        words = ['adj', 'adjust', 'adjustment', 'business', 'cancel', 'CEO', 'CFO', 'charity', 'clear', 'corr',
                 'correct', 'correction', 'development', 'director', 'directors', 'error', 'fix', 'fraud', 'gift',
                 'help', 'impair', 'impairment', 'lobby', 'lobbying', 'mistake', 'per', 'restate', 'return', 'reverse',
                 'suspense', 'variance', 'wash off', 'write down', 'write off', 'write-off', 'writeoff']
        self.sus_lwidget_defined.addItems(words)

    def move_word(self):
        res = []
        selection = self.sus_lwidget_defined.selectionModel()
        indexes = sorted(selection.selectedIndexes(), key=lambda i: i.row())
        for index in indexes:
            print(index.row(), index.data())
            res.append(index.data())
            # if index.data() not in self.suspicious_words:
            #     self.suspicious_words.append(index.data())
            self.sus_lwidget_defined.takeItem(index.row())

        self.sus_lwidget.addItems(res)

    def move_back_word(self):
        res = []
        selection = self.sus_lwidget.selectionModel()
        indexes = sorted(selection.selectedIndexes(), key=lambda i: i.row())
        for index in indexes:
            print(index.row(), index.data())
            res.append(index.data())
            # self.suspicious_words.pop(self.suspicious_words.index(index.data()))
            self.sus_lwidget.takeItem(index.row())

        self.sus_lwidget_defined.addItems(res)

    def add_words(self):
        x = self.sus_ledit.text()
        print(x)
        if x:
            self.sus_lwidget.addItem(x)
            self.sus_ledit.clear()

    def remove_words(self):
        word = self.sus_lwidget.currentItem().text()
        index = self.sus_lwidget.currentRow()
        self.sus_lwidget.takeItem(index)

    def clear_words(self):
        self.sus_lwidget.clear()

    def add_pattern(self):
        only_double = QDoubleValidator()
        self.wh_ln.setValidator(only_double)
        x = self.wh_ln.text()
        if x:
            self.wh_lwidget.addItem(x)
            self.wh_ln.clear()

    def remove_pattern(self):
        index = self.wh_lwidget.currentRow()
        self.wh_lwidget.takeItem(index)

    def clear_patterns(self):
        self.wh_lwidget.clear()

    def add_estimates(self):
        reg_ex = QRegExp("^[0-9]*$")
        input_validator = QRegExpValidator(reg_ex, self.ac_ln)
        self.ac_ln.setValidator(input_validator)
        x = self.ac_ln.text()
        if x:
            self.ac_lwidget.addItem(x)
            self.ac_ln.clear()

    def remove_estimates(self):
        index = self.ac_lwidget.currentRow()
        self.ac_lwidget.takeItem(index)

    def clear_estimates(self):
        self.ac_lwidget.clear()

    def initial_testing(self):
        self.dataframe = get_dataframe(name='conv_GL.csv')
        # self.dataframe = init_date(self.dataframe)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            print("esc")
            self.close()

    def refresh_names(self):
        created = self.cmb_user_analysis.currentText()
        names = self.dataframe.groupby([created])
        names = list(names.groups.keys())
        print(names)
        self.user_listwidget.clear()
        for i in names:
            item = QListWidgetItem(str(i))
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            item.setFont(QFont("Arial", 14))
            self.user_listwidget.addItem(item)

    def init_duplicate_listwidget(self):
        self.duplicate_listwidget.clear()
        for i in self.dataframe.columns.values.tolist():
            item = QListWidgetItem(i)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            item.setFont(QFont("Arial", 14))
            self.duplicate_listwidget.addItem(item)
